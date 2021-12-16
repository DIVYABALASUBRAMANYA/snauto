import xlrd
import openpyxl


def read_data(sheet_name):
    listli = []
    path = "/Users/dbalasub/PycharmProjects/pythonProject/DataForTest/DataSheet.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows + 1):
        row = []
        for c in range(1, cols+1):
            row.append(sheet.cell(r, c).value)
        listli.append(row)
        # print(val)
        # listli.append(sheet.cell(r, c).value)
    return listli


















